from openpyxl import Workbook
from openpyxl.styles import PatternFill


def generate_excel(results, output_file):
    """
    Generates a formatted Excel file with colored rows according to changes:
    - Red    = Removed
    - Yellow = Modified
    - Green  = Added

    results is a list of lists:
    [ConfigMap, Key, Before, After, Status, BeforeTime, AfterTime]
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Diff Report"

    # Header row
    ws.append(["ConfigMap", "Key", "Before", "After", "Status", "BeforeTime", "AfterTime"])

    # Define color styles
    fill_removed  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_modified = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_added    = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for row in results:
        ws.append(row)
        status = row[4]  # Status column
        excel_row = ws.max_row

        # Apply coloring based on status
        if status == "Removed":
            for col in range(1, 8):
                ws.cell(row=excel_row, column=col).fill = fill_removed

        elif status == "Modified":
            for col in range(1, 8):
                ws.cell(row=excel_row, column=col).fill = fill_modified

        elif status == "Added":
            for col in range(1, 8):
                ws.cell(row=excel_row, column=col).fill = fill_added

    wb.save(output_file)
    print(f"✔ Excel saved → {output_file}")
